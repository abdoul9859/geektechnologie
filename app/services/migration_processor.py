import asyncio
import json
import csv
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional
import threading
import time
import requests
import hashlib
import os

from ..database import Migration, MigrationLog, Product, Client, Supplier, ProductVariant, StockMovement, get_next_id
from ..routers.cache import set_cache_item


class MigrationProcessor:
    """Service de traitement des migrations en arriere-plan"""

    def __init__(self):
        self.running_migrations: Dict[int, bool] = {}
        self.processing_thread = None
        self.should_stop = False

    def start_background_processor(self):
        """Demarre le processeur en arriere-plan"""
        if self.processing_thread is None or not self.processing_thread.is_alive():
            self.should_stop = False
            self.processing_thread = threading.Thread(target=self._background_worker, daemon=True)
            self.processing_thread.start()
            print("Processeur de migrations demarre")

    def stop_background_processor(self):
        """Arrete le processeur en arriere-plan"""
        self.should_stop = True
        if self.processing_thread and self.processing_thread.is_alive():
            self.processing_thread.join(timeout=5)
            print("Processeur de migrations arrete")

    def _background_worker(self):
        """Worker en arriere-plan qui traite les migrations"""
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            while not self.should_stop:
                try:
                    loop.run_until_complete(self._async_check_pending())
                except Exception as e:
                    print(f"Erreur dans le worker de migrations: {e}")

                time.sleep(2)  # Verifier toutes les 2 secondes
        finally:
            loop.close()

    async def _async_check_pending(self):
        """Cherche les migrations en attente et lance leur traitement"""
        # Chercher les migrations en attente de traitement
        pending_migrations = await Migration.find(
            {"status": "running", "migration_id": {"$nin": list(self.running_migrations.keys())}}
        ).to_list()

        for migration in pending_migrations:
            if migration.migration_id not in self.running_migrations:
                # Marquer comme en cours de traitement
                self.running_migrations[migration.migration_id] = True

                # Traiter la migration dans un thread separe
                thread = threading.Thread(
                    target=self._process_migration,
                    args=(migration.migration_id,),
                    daemon=True
                )
                thread.start()

    def _process_migration(self, migration_id: int):
        """Traite une migration specifique (sync wrapper pour thread)"""
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            loop.run_until_complete(self._async_process_migration(migration_id))
        finally:
            loop.close()

    async def _async_process_migration(self, migration_id: int):
        """Traite une migration specifique (async)"""
        try:
            migration = await Migration.find_one(Migration.migration_id == migration_id)
            if not migration:
                return

            await self._add_log(migration_id, "info", f"Debut du traitement de la migration: {migration.name}")

            # Verifier si un fichier est associe
            if migration.file_name:
                file_path = Path("uploads") / "migrations" / migration.file_name
                if file_path.exists():
                    await self._add_log(migration_id, "info", f"Traitement du fichier: {migration.file_name}")

                    # Traiter selon le type de migration
                    success = await self._process_file(migration, file_path)

                    if success:
                        # Marquer comme terminee avec succes
                        migration.status = "completed"
                        migration.completed_at = datetime.utcnow()
                        await self._add_log(migration_id, "success", f"Migration terminee avec succes. {migration.success_records} enregistrements traites.")
                    else:
                        # Marquer comme echouee
                        migration.status = "failed"
                        migration.completed_at = datetime.utcnow()
                        migration.error_message = "Erreur lors du traitement du fichier"
                        await self._add_log(migration_id, "error", "Migration echouee lors du traitement du fichier")
                else:
                    # Fichier non trouve
                    migration.status = "failed"
                    migration.completed_at = datetime.utcnow()
                    migration.error_message = "Fichier non trouve"
                    await self._add_log(migration_id, "error", f"Fichier non trouve: {migration.file_name}")
            else:
                # Pas de fichier - migration de test
                await self._simulate_processing(migration)

            await migration.save()

        except Exception as e:
            print(f"Erreur lors du traitement de la migration {migration_id}: {e}")
            migration = await Migration.find_one(Migration.migration_id == migration_id)
            if migration:
                migration.status = "failed"
                migration.completed_at = datetime.utcnow()
                migration.error_message = str(e)
                await self._add_log(migration_id, "error", f"Erreur critique: {str(e)}")
                await migration.save()

        finally:
            # Retirer de la liste des migrations en cours
            if migration_id in self.running_migrations:
                del self.running_migrations[migration_id]

    async def _process_file(self, migration: Migration, file_path: Path) -> bool:
        """Traite un fichier de migration selon son type"""
        try:
            file_extension = file_path.suffix.lower()

            if file_extension == '.csv':
                return await self._process_csv_file(migration, file_path)
            elif file_extension in ['.xlsx', '.xls']:
                return await self._process_excel_file(migration, file_path)
            elif file_extension == '.json':
                return await self._process_json_file(migration, file_path)
            else:
                await self._add_log(migration.migration_id, "error", f"Format de fichier non supporte: {file_extension}")
                return False

        except Exception as e:
            await self._add_log(migration.migration_id, "error", f"Erreur lors du traitement du fichier: {str(e)}")
            return False

    async def _process_csv_file(self, migration: Migration, file_path: Path) -> bool:
        """Traite un fichier CSV"""
        try:
            with open(file_path, 'r', encoding='utf-8', newline='') as csvfile:
                # Detecter le delimiteur
                sample = csvfile.read(1024)
                csvfile.seek(0)
                sniffer = csv.Sniffer()
                delimiter = sniffer.sniff(sample).delimiter

                reader = csv.DictReader(csvfile, delimiter=delimiter)
                rows = list(reader)
                migration.total_records = len(rows)

                await self._add_log(migration.migration_id, "info", f"Fichier CSV charge: {migration.total_records} lignes")

                success_count = 0
                error_count = 0

                for index, row in enumerate(rows):
                    try:
                        # Traiter selon le type de migration
                        if migration.type == "products":
                            success = await self._import_product_from_row(row)
                        elif migration.type == "clients":
                            success = await self._import_client_from_row(row)
                        elif migration.type == "suppliers":
                            success = await self._import_supplier_from_row(row)
                        else:
                            success = True  # Migration generique

                        if success:
                            success_count += 1
                        else:
                            error_count += 1

                        # Mettre a jour les compteurs periodiquement
                        if (index + 1) % 10 == 0:
                            migration.processed_records = index + 1
                            migration.success_records = success_count
                            migration.error_records = error_count
                            await migration.save()

                            await self._add_log(migration.migration_id, "info", f"Progression: {index + 1}/{migration.total_records} lignes traitees")

                    except Exception as e:
                        error_count += 1
                        await self._add_log(migration.migration_id, "warning", f"Erreur ligne {index + 1}: {str(e)}")

                # Mise a jour finale
                migration.processed_records = migration.total_records
                migration.success_records = success_count
                migration.error_records = error_count

                return error_count == 0 or success_count > 0

        except Exception as e:
            await self._add_log(migration.migration_id, "error", f"Erreur lors de la lecture du CSV: {str(e)}")
            return False

    async def _process_excel_file(self, migration: Migration, file_path: Path) -> bool:
        """Traite un fichier Excel (necessite openpyxl)"""
        try:
            import openpyxl
            from openpyxl import load_workbook

            # Charger le fichier Excel
            workbook = load_workbook(file_path, read_only=True)
            worksheet = workbook.active

            # Lire les en-tetes
            headers = []
            for cell in worksheet[1]:
                if cell.value:
                    headers.append(str(cell.value).strip().lower())
                else:
                    headers.append(None)

            await self._add_log(migration.migration_id, "info", f"En-tetes detectes: {headers}")

            # Compter les lignes de donnees
            total_rows = 0
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    total_rows += 1

            migration.total_records = total_rows
            await self._add_log(migration.migration_id, "info", f"Fichier Excel charge: {total_rows} lignes de donnees")

            if total_rows == 0:
                await self._add_log(migration.migration_id, "warning", "Aucune donnee trouvee dans le fichier Excel")
                return False

            success_count = 0
            error_count = 0

            # Traiter chaque ligne
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(cell is not None for cell in row):
                    continue

                try:
                    # Creer un dictionnaire avec les donnees de la ligne
                    row_data = {}
                    for i, value in enumerate(row):
                        if i < len(headers) and headers[i]:
                            row_data[headers[i]] = value

                    # Traiter selon le type de migration
                    if migration.type == "products":
                        success = await self._import_product_from_excel_row(row_data)
                    elif migration.type == "clients":
                        success = await self._import_client_from_excel_row(row_data)
                    elif migration.type == "suppliers":
                        success = await self._import_supplier_from_excel_row(row_data)
                    else:
                        success = True  # Migration generique

                    if success:
                        success_count += 1
                    else:
                        error_count += 1

                    # Mettre a jour les compteurs periodiquement
                    if (row_num - 1) % 10 == 0:
                        migration.processed_records = row_num - 1
                        migration.success_records = success_count
                        migration.error_records = error_count
                        await migration.save()

                        await self._add_log(migration.migration_id, "info", f"Progression: {row_num - 1}/{total_rows} lignes traitees")

                except Exception as e:
                    error_count += 1
                    error_msg = f"Erreur ligne {row_num}: {str(e)}"
                    await self._add_log(migration.migration_id, "error", error_msg)

            # Mise a jour finale
            migration.processed_records = total_rows
            migration.success_records = success_count
            migration.error_records = error_count

            workbook.close()
            return error_count == 0 or success_count > 0

        except ImportError:
            await self._add_log(migration.migration_id, "error", "Bibliotheque openpyxl non disponible - installez avec: pip install openpyxl")
            return False
        except Exception as e:
            await self._add_log(migration.migration_id, "error", f"Erreur lors de la lecture du fichier Excel: {str(e)}")
            return False

    async def _process_json_file(self, migration: Migration, file_path: Path) -> bool:
        """Traite un fichier JSON"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            if isinstance(data, list):
                migration.total_records = len(data)
                await self._add_log(migration.migration_id, "info", f"Fichier JSON charge: {migration.total_records} enregistrements")

                success_count = 0
                error_count = 0

                for index, item in enumerate(data):
                    try:
                        # Traiter selon le type
                        if migration.type == "products":
                            success = await self._import_product_from_dict(item)
                        elif migration.type == "clients":
                            success = await self._import_client_from_dict(item)
                        elif migration.type == "suppliers":
                            success = await self._import_supplier_from_dict(item)
                        else:
                            success = True

                        if success:
                            success_count += 1
                        else:
                            error_count += 1

                    except Exception as e:
                        error_count += 1
                        await self._add_log(migration.migration_id, "warning", f"Erreur enregistrement {index + 1}: {str(e)}")

                migration.processed_records = migration.total_records
                migration.success_records = success_count
                migration.error_records = error_count

                return error_count == 0 or success_count > 0
            else:
                await self._add_log(migration.migration_id, "error", "Le fichier JSON doit contenir un tableau")
                return False

        except Exception as e:
            await self._add_log(migration.migration_id, "error", f"Erreur lors de la lecture du JSON: {str(e)}")
            return False

    async def _simulate_processing(self, migration: Migration):
        """Simule le traitement d'une migration sans fichier"""
        migration.total_records = 100

        await self._add_log(migration.migration_id, "info", "Debut de la simulation de traitement")

        for i in range(0, 101, 10):
            migration.processed_records = i
            migration.success_records = i - (i // 20)  # Quelques erreurs simulees
            migration.error_records = i // 20

            await migration.save()

            await self._add_log(migration.migration_id, "info", f"Progression: {i}/100 enregistrements traites")
            time.sleep(1)  # Simuler le temps de traitement

        migration.status = "completed"
        migration.completed_at = datetime.utcnow()

    async def _import_product_from_row(self, row) -> bool:
        """Importe un produit depuis une ligne CSV/Excel"""
        try:
            from decimal import Decimal

            new_id = await get_next_id("products")
            product = Product(
                product_id=new_id,
                name=str(row.get('name', row.get('nom', ''))),
                description=str(row.get('description', '')),
                price=Decimal(str(row.get('price', row.get('prix', 0)))),
                quantity=int(row.get('stock', row.get('quantite', 0))),
                category=str(row.get('category', row.get('categorie', ''))),
            )
            await product.insert()
            return True
        except Exception:
            return False

    async def _import_product_from_excel_row(self, row_data: dict) -> bool:
        """Importe un produit depuis une ligne Excel avec structure complete"""
        try:
            from decimal import Decimal
            from datetime import datetime

            # Extraire les donnees avec des noms de colonnes flexibles
            name = self._get_value(row_data, ['name', 'nom', 'product_name', 'produit'])
            if not name:
                return False

            description = self._get_value(row_data, ['description', 'desc', 'description_produit'])
            price = self._get_float_value(row_data, ['price', 'prix', 'unit_price', 'prix_unitaire'])
            purchase_price = self._get_float_value(row_data, ['purchase_price', 'prix_achat', 'cost', 'cout'])
            quantity = self._get_int_value(row_data, ['quantity', 'quantite', 'quantite', 'stock', 'qty'])
            category = self._get_value(row_data, ['category', 'categorie', 'categorie', 'cat'])
            brand = self._get_value(row_data, ['brand', 'marque', 'fabricant'])
            model = self._get_value(row_data, ['model', 'modele', 'modele', 'reference'])
            barcode = self._get_value(row_data, ['barcode', 'code_barre', 'code-barres', 'ean', 'sku'])
            condition = self._get_value(row_data, ['condition', 'etat', 'state'])
            notes = self._get_value(row_data, ['notes', 'commentaires', 'remarques'])
            image_url = self._get_value(row_data, ['image_path', 'image', 'photo', 'picture', 'img', 'image_url', 'url_image'])

            # Si une URL d'image est fournie, la telecharger
            image_path = None
            if image_url:
                image_path = self._download_and_save_image(image_url, name)

            # Detecter si c'est un produit avec variantes
            # Chercher des colonnes de variantes (IMEI, serie, etc.)
            imei_serial = self._get_value(row_data, ['imei', 'serial', 'imei_serial', 'numero_serie', 'numero_serie'])
            variant_barcode = self._get_value(row_data, ['variant_barcode', 'code_barre_variante', 'barcode_variant'])
            variant_condition = self._get_value(row_data, ['variant_condition', 'condition_variante', 'etat_variante'])

            has_variants = bool(imei_serial)  # Si IMEI fourni, c'est une variante

            # Valeurs par defaut
            if not price:
                price = 0.0
            if not purchase_price:
                purchase_price = price
            if not quantity:
                quantity = 0
            if not condition:
                condition = "neuf"
            if not variant_condition:
                variant_condition = condition

            if has_variants:
                # Produit avec variantes
                return await self._import_product_with_variants(
                    name, description, price, purchase_price, category, brand, model,
                    condition, notes, imei_serial, variant_barcode, variant_condition, image_path
                )
            else:
                # Produit sans variantes
                return await self._import_simple_product(
                    name, description, price, purchase_price, quantity, category,
                    brand, model, barcode, condition, notes, image_path
                )

        except Exception as e:
            return False

    async def _import_simple_product(self, name: str, description: str, price: float,
                              purchase_price: float, quantity: int, category: str, brand: str,
                              model: str, barcode: str, condition: str, notes: str, image_path: str = None) -> bool:
        """Importe un produit simple sans variantes"""
        try:
            from decimal import Decimal
            from datetime import datetime

            new_id = await get_next_id("products")

            # Creer le produit
            product = Product(
                product_id=new_id,
                name=name,
                description=description or "",
                price=Decimal(str(price)),
                purchase_price=Decimal(str(purchase_price)),
                quantity=quantity,
                category=category or "",
                brand=brand or "",
                model=model or "",
                barcode=barcode if barcode else None,
                condition=condition.lower(),
                has_unique_serial=False,
                entry_date=datetime.now(),
                notes=notes or "",
                image_path=image_path if image_path else None
            )

            await product.insert()

            # Creer un mouvement de stock d'entree si quantite > 0
            if quantity > 0:
                movement_id = await get_next_id("stock_movements")
                stock_movement = StockMovement(
                    movement_id=movement_id,
                    product_id=product.product_id,
                    quantity=quantity,
                    movement_type="IN",
                    reference_type="IMPORT_EXCEL",
                    reference_id=None,
                    notes="Import depuis fichier Excel",
                    unit_price=Decimal(str(price))
                )
                await stock_movement.insert()

            return True

        except Exception as e:
            return False

    async def _import_product_with_variants(self, name: str, description: str, price: float,
                                     purchase_price: float, category: str, brand: str, model: str,
                                     condition: str, notes: str, imei_serial: str, variant_barcode: str,
                                     variant_condition: str, image_path: str = None) -> bool:
        """Importe un produit avec variantes"""
        try:
            from decimal import Decimal
            from datetime import datetime

            new_id = await get_next_id("products")

            # Creer le produit (sans code-barres car il aura des variantes)
            product = Product(
                product_id=new_id,
                name=name,
                description=description or "",
                price=Decimal(str(price)),
                purchase_price=Decimal(str(purchase_price)),
                quantity=1,  # 1 variante
                category=category or "",
                brand=brand or "",
                model=model or "",
                barcode=None,  # Pas de code-barres au niveau produit
                condition=condition.lower(),
                has_unique_serial=True,  # Produit avec variantes
                entry_date=datetime.now(),
                notes=notes or "",
                image_path=image_path if image_path else None
            )

            await product.insert()

            # Creer la variante
            variant_id = await get_next_id("product_variants")
            variant = ProductVariant(
                variant_id=variant_id,
                product_id=product.product_id,
                imei_serial=imei_serial,
                barcode=variant_barcode if variant_barcode else None,
                condition=variant_condition.lower(),
                is_sold=False
            )

            await variant.insert()

            # Creer un mouvement de stock d'entree
            movement_id = await get_next_id("stock_movements")
            stock_movement = StockMovement(
                movement_id=movement_id,
                product_id=product.product_id,
                quantity=1,
                movement_type="IN",
                reference_type="IMPORT_EXCEL",
                reference_id=None,
                notes="Import variante depuis fichier Excel",
                unit_price=Decimal(str(price))
            )
            await stock_movement.insert()

            return True

        except Exception as e:
            return False

    def _get_value(self, row_data: dict, possible_keys: list) -> str:
        """Recupere une valeur en testant plusieurs cles possibles"""
        for key in possible_keys:
            # Chercher la cle exacte
            value = row_data.get(key)
            if value is not None and str(value).strip():
                return str(value).strip()

            # Chercher des variantes avec espaces et accents
            for row_key in row_data.keys():
                if key.lower() in row_key.lower() or row_key.lower() in key.lower():
                    value = row_data.get(row_key)
                    if value is not None and str(value).strip():
                        return str(value).strip()
        return ""

    def _get_float_value(self, row_data: dict, possible_keys: list) -> float:
        """Recupere une valeur float en testant plusieurs cles possibles"""
        for key in possible_keys:
            # Chercher la cle exacte
            value = row_data.get(key)
            if value is not None:
                try:
                    return float(value)
                except (ValueError, TypeError):
                    pass

            # Chercher des variantes avec espaces et accents
            for row_key in row_data.keys():
                if key.lower() in row_key.lower() or row_key.lower() in key.lower():
                    value = row_data.get(row_key)
                    if value is not None:
                        try:
                            return float(value)
                        except (ValueError, TypeError):
                            pass
        return 0.0

    def _download_and_save_image(self, image_url: str, product_name: str) -> Optional[str]:
        """Telecharge une image depuis une URL et la sauvegarde localement"""
        try:
            # Verifier si c'est une URL valide
            if not image_url.startswith(('http://', 'https://')):
                # Si ce n'est pas une URL, considerer que c'est deja un chemin local
                return image_url

            # Creer le dossier de destination s'il n'existe pas
            upload_dir = Path("static/uploads/products")
            upload_dir.mkdir(parents=True, exist_ok=True)

            # Telecharger l'image
            response = requests.get(image_url, timeout=10, stream=True)
            response.raise_for_status()

            # Determiner l'extension du fichier
            content_type = response.headers.get('content-type', '')
            extension = '.jpg'  # Par defaut
            if 'png' in content_type:
                extension = '.png'
            elif 'jpeg' in content_type or 'jpg' in content_type:
                extension = '.jpg'
            elif 'webp' in content_type:
                extension = '.webp'
            elif 'gif' in content_type:
                extension = '.gif'

            # Generer un nom de fichier unique base sur le nom du produit et un hash
            safe_name = "".join(c for c in product_name if c.isalnum() or c in (' ', '-', '_')).strip()
            safe_name = safe_name.replace(' ', '_')[:50]  # Limiter la longueur
            timestamp = int(datetime.now().timestamp())
            url_hash = hashlib.md5(image_url.encode()).hexdigest()[:8]
            filename = f"{safe_name}_{timestamp}_{url_hash}{extension}"

            # Chemin complet du fichier
            file_path = upload_dir / filename

            # Sauvegarder l'image
            with open(file_path, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)

            # Retourner le chemin relatif pour la base de donnees
            return f"static/uploads/products/{filename}"

        except requests.exceptions.RequestException as e:
            print(f"Erreur lors du telechargement de l'image {image_url}: {e}")
            return None
        except Exception as e:
            print(f"Erreur lors de la sauvegarde de l'image: {e}")
            return None

    def _get_int_value(self, row_data: dict, possible_keys: list) -> int:
        """Recupere une valeur int en testant plusieurs cles possibles"""
        for key in possible_keys:
            # Chercher la cle exacte
            value = row_data.get(key)
            if value is not None:
                try:
                    return int(float(value))  # Convertir via float pour gerer "1.0"
                except (ValueError, TypeError):
                    pass

            # Chercher des variantes avec espaces et accents
            for row_key in row_data.keys():
                if key.lower() in row_key.lower() or row_key.lower() in key.lower():
                    value = row_data.get(row_key)
                    if value is not None:
                        try:
                            return int(float(value))  # Convertir via float pour gerer "1.0"
                        except (ValueError, TypeError):
                            pass
        return 0

    async def _import_client_from_row(self, row) -> bool:
        """Importe un client depuis une ligne CSV/Excel"""
        try:
            new_id = await get_next_id("clients")
            client = Client(
                client_id=new_id,
                name=str(row.get('name', row.get('nom', ''))),
                email=str(row.get('email', '')),
                phone=str(row.get('phone', row.get('telephone', ''))),
                address=str(row.get('address', row.get('adresse', '')))
            )
            await client.insert()
            return True
        except Exception:
            return False

    async def _import_client_from_excel_row(self, row_data: dict) -> bool:
        """Importe un client depuis une ligne Excel"""
        return await self._import_client_from_row(row_data)

    async def _import_supplier_from_row(self, row) -> bool:
        """Importe un fournisseur depuis une ligne CSV/Excel"""
        try:
            new_id = await get_next_id("suppliers")
            supplier = Supplier(
                supplier_id=new_id,
                name=str(row.get('name', row.get('nom', ''))),
                email=str(row.get('email', '')),
                phone=str(row.get('phone', row.get('telephone', ''))),
                address=str(row.get('address', row.get('adresse', '')))
            )
            await supplier.insert()
            return True
        except Exception:
            return False

    async def _import_supplier_from_excel_row(self, row_data: dict) -> bool:
        """Importe un fournisseur depuis une ligne Excel"""
        return await self._import_supplier_from_row(row_data)

    async def _import_product_from_dict(self, data: dict) -> bool:
        """Importe un produit depuis un dictionnaire JSON"""
        return await self._import_product_from_row(data)

    async def _import_client_from_dict(self, data: dict) -> bool:
        """Importe un client depuis un dictionnaire JSON"""
        return await self._import_client_from_row(data)

    async def _import_supplier_from_dict(self, data: dict) -> bool:
        """Importe un fournisseur depuis un dictionnaire JSON"""
        return await self._import_supplier_from_row(data)

    async def _add_log(self, migration_id: int, level: str, message: str):
        """Ajoute un log a une migration"""
        try:
            log_id = await get_next_id("migration_logs")
            log = MigrationLog(
                log_id=log_id,
                migration_id=migration_id,
                level=level,
                message=message,
                timestamp=datetime.utcnow()
            )
            await log.insert()

            # Mettre en cache pour les performances
            cache_key = f"migration_logs:{migration_id}"
            set_cache_item(cache_key, {"last_log": message, "level": level}, ttl_hours=1, cache_type="migration")

        except Exception as e:
            print(f"Erreur lors de l'ajout du log: {e}")

# Instance globale du processeur
migration_processor = MigrationProcessor()
